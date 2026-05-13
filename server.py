"""
StockMonitor · Flask 后端
────────────────────────
启动方式：
    python server.py
然后浏览器打开 http://localhost:5000

API 路由：
    GET  /api/stocks          → 所有股票摘要列表
    GET  /api/stock/<code>    → 单只股票完整数据
    POST /api/stock/add       → 添加股票  body: {"code":"600519"}
    POST /api/stock/remove    → 删除股票  body: {"code":"600519"}
    POST /api/stock/refresh   → 刷新单只  body: {"code":"600519"}
    GET  /api/status          → 服务状态
"""

import os
import sys
import json
import time
import threading
import datetime
import traceback
from pathlib import Path
from collections import deque
from concurrent.futures import ThreadPoolExecutor

from flask import Flask, jsonify, request, send_from_directory

# ── 项目模块 ───────────────────────────────────────────────────
try:
    from config import Config
    from fetcher import fetch_kline, fetch_stock_info
    from indicators import calc_indicators
    from analysis import find_support_resistance, describe_sr_relation
    from ai_advisor import build_prompt, AIAdvisor
    from decision import make_final_decision, rule_engine
    HAS_MODULES = True
except ImportError as e:
    print(f"[警告] 部分模块未安装，将使用 Demo 数据: {e}")
    HAS_MODULES = False

# ── 大宗商品模块 ───────────────────────────────────────────────
try:
    from commodity_fetcher import commodity_fetcher
    from commodity_ai import CommodityAIAdvisor
    HAS_COMMODITY = True
except ImportError as e:
    print(f"[警告] 大宗商品模块未找到: {e}")
    HAS_COMMODITY = False

# ── 板块/资金/日历模块（可选）──────────────────────────────────
try:
    from sector_fetcher   import get_sector_info
    from capital_fetcher  import get_northbound, get_margin
    from calendar_fetcher import get_upcoming_events
    HAS_EXTRA = True
except ImportError as e:
    print(f"[警告] 扩展模块未找到: {e}")
    HAS_EXTRA = False

# ── 日志初始化（必须在其他 logger 调用前执行）────────────────────
# 没有这个调用的话，loguru 会用默认 handler（DEBUG 级别全开）刷屏控制台
try:
    from logger_setup import setup_logger
    _APP_DIR = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
    setup_logger(str(_APP_DIR))
except Exception as _log_err:
    print(f"[警告] 日志模块初始化失败: {_log_err}")

# 大宗商品全局状态
_commodity_analysis: dict = {}
_analysis_lock = threading.Lock()
_analysis_loading = threading.Event()

# ── 持久化文件路径 ─────────────────────────────────────────────
WATCHLIST_FILE     = Path("watchlist.json")
ANALYSIS_CACHE_FILE = Path("analysis_cache.json")
ALERTS_FILE        = Path("alerts.json")

# ── 大宗商品价格历史（用于 Sparkline）─────────────────────────
# {code: deque([p1,p2,...], maxlen=24)}  每5分钟一个点，最多2小时
_price_history: dict = {}
_history_lock = threading.Lock()

# ── 价格预警 ───────────────────────────────────────────────────
# [{id,type,code,name,direction:'above'|'below',threshold,triggered,triggered_at}]
_alerts: list = []
_alerts_lock = threading.Lock()
_triggered_queue: list = []   # 已触发待通知列表


# ══════════════════════════════════════════════════════════════
#  持久化工具函数
# ══════════════════════════════════════════════════════════════

def _load_watchlist(default_codes: list) -> list:
    """从 watchlist.json 加载自选股列表，不存在则用 .env 默认值"""
    try:
        if WATCHLIST_FILE.exists():
            data = json.loads(WATCHLIST_FILE.read_text(encoding="utf-8"))
            codes = [c.strip() for c in data.get("codes", []) if c.strip()]
            if codes:
                print(f"[自选股] 从 watchlist.json 加载 {len(codes)} 只")
                return codes
    except Exception as e:
        print(f"[自选股] 读取失败，使用默认: {e}")
    return default_codes

def _save_watchlist(codes: list):
    try:
        WATCHLIST_FILE.write_text(
            json.dumps({"codes": codes, "updated_at": datetime.datetime.now().isoformat()},
                       ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    except Exception as e:
        print(f"[自选股] 保存失败: {e}")

def _load_analysis_cache() -> dict:
    """加载大宗商品 AI 分析缓存（24小时内有效）"""
    try:
        if ANALYSIS_CACHE_FILE.exists():
            data = json.loads(ANALYSIS_CACHE_FILE.read_text(encoding="utf-8"))
            saved = datetime.datetime.fromisoformat(data.get("saved_at", "2000-01-01"))
            age_h = (datetime.datetime.now() - saved).total_seconds() / 3600
            if age_h < 24:
                print(f"[缓存] 加载 AI 分析缓存（{age_h:.1f}小时前）")
                return data.get("analysis", {})
    except Exception as e:
        print(f"[缓存] 读取失败: {e}")
    return {}

def _save_analysis_cache(analysis: dict):
    try:
        ANALYSIS_CACHE_FILE.write_text(
            json.dumps({"saved_at": datetime.datetime.now().isoformat(), "analysis": analysis},
                       ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    except Exception as e:
        print(f"[缓存] 保存失败: {e}")

def _load_alerts():
    global _alerts
    try:
        if ALERTS_FILE.exists():
            _alerts = json.loads(ALERTS_FILE.read_text(encoding="utf-8"))
            print(f"[预警] 加载 {len(_alerts)} 条预警规则")
    except Exception as e:
        print(f"[预警] 加载失败: {e}")
        _alerts = []

def _save_alerts():
    try:
        with _alerts_lock:
            data = list(_alerts)
        ALERTS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"[预警] 保存失败: {e}")

def _check_alerts(code: str, name: str, price: float):
    """检查价格预警，触发时加入通知队列"""
    with _alerts_lock:
        for alert in _alerts:
            if alert.get("code") != code or alert.get("triggered"):
                continue
            thr = float(alert.get("threshold", 0))
            direction = alert.get("direction", "above")
            hit = (direction == "above" and price >= thr) or \
                  (direction == "below" and price <= thr)
            if hit:
                alert["triggered"] = True
                alert["triggered_at"] = datetime.datetime.now().isoformat()
                payload = {
                    "code": code, "name": name, "price": price,
                    "direction": direction, "threshold": thr,
                    "triggered_at": alert["triggered_at"],
                }
                _triggered_queue.append(payload)
                # WebSocket 即时推送（替代 30 秒轮询，延迟 ~30s → ~瞬时）
                _ws_emit("alert_triggered", payload)
                print(f"[预警] ★ {name}({code}) 价格 {price} {direction} {thr}")
    _save_alerts()

# ══════════════════════════════════════════════════════════════
#  Demo 数据（无依赖时使用）
# ══════════════════════════════════════════════════════════════
DEMO_STOCKS = {
    "600519": {
        "code": "600519", "name": "贵州茅台", "price": 1567.00, "change": 36.60,
        "decision": "强烈买入", "rule": "买入",
        "dif": 12.45, "dea": 8.32, "hist": 8.26,
        "bbU": 1602, "bbM": 1548, "bbL": 1494, "bbPct": 0.64, "bbW": 6.97,
        "rsi": 58.4, "vr": 1.82, "tr": 1.23,
        "ma5": 1552, "ma10": 1538, "ma20": 1548, "ma60": 1490,
        "sups": [1520, 1490, 1455], "ress": [1600, 1650, 1700],
        "aiDecision": "强烈买入", "aiConf": "高",
        "aiSum": "MACD金叉形成，DIF(12.45)站稳零轴上方，配合布林带中轨之上运行，带口向上扩展。量比1.82明显放量，多方力度占优。最近支撑1520元较扎实（当前价高于支撑+3.1%），综合技术面偏强。",
        "aiR": [
            "MACD DIF(12.45) > DEA(8.32)，金叉持续有效，柱线走强",
            "价格(1567)站布林中轨(1548)，位于带内64%，趋势健康",
            "量比1.82，放量支撑上涨，无量价背离信号",
            "RSI=58.4，尚未进入超买区域，上升空间充足",
            "均线多头排列：MA5(1552)>MA20(1548)>MA60(1490)"
        ],
        "aiE": 1555, "aiS": 1490, "aiT": 1680,
        "macdD": [2.1, 3.4, 5.2, 4.8, 7.2, 9.1, 10.3, 11.8, 12.4, 12.45],
        "deaD":  [1.2, 1.8, 2.9, 3.5, 4.8, 5.9, 7.1, 7.9, 8.2, 8.32],
        "priceD": [1510, 1522, 1498, 1515, 1530, 1545, 1552, 1558, 1562, 1567],
        "lbls": ["5/20","5/21","5/22","5/23","5/24","5/27","5/28","5/29","5/30","5/31"],
        "updatedAt": datetime.datetime.now().isoformat(),
    },
    "000001": {
        "code": "000001", "name": "平安银行", "price": 11.23, "change": -0.18,
        "decision": "观望", "rule": "观望",
        "dif": -0.12, "dea": 0.08, "hist": -0.40,
        "bbU": 11.85, "bbM": 11.45, "bbL": 11.05, "bbPct": 0.48, "bbW": 7.25,
        "rsi": 44.2, "vr": 0.72, "tr": 0.58,
        "ma5": 11.35, "ma10": 11.42, "ma20": 11.45, "ma60": 11.80,
        "sups": [11.05, 10.80, 10.50], "ress": [11.45, 11.85, 12.20],
        "aiDecision": "观望", "aiConf": "中",
        "aiSum": "MACD于零轴下方死叉，价格处于布林带中轨下方，量比0.72属缩量状态。均线呈空头排列，短期下行压力较大，建议耐心等待底部信号确认。",
        "aiR": [
            "MACD死叉且DIF(-0.12)在零轴下方，空头占优",
            "均线空头排列，MA5(11.35)<MA20(11.45)<MA60(11.80)",
            "缩量下跌，动能不足，反弹力度存疑"
        ],
        "aiE": None, "aiS": None, "aiT": None,
        "macdD": [0.5, 0.3, 0.1, -0.05, -0.08, -0.1, -0.11, -0.12, -0.12, -0.12],
        "deaD":  [0.4, 0.35, 0.28, 0.2, 0.15, 0.12, 0.10, 0.09, 0.08, 0.08],
        "priceD": [11.52, 11.48, 11.45, 11.40, 11.38, 11.32, 11.28, 11.25, 11.24, 11.23],
        "lbls": ["5/20","5/21","5/22","5/23","5/24","5/27","5/28","5/29","5/30","5/31"],
        "updatedAt": datetime.datetime.now().isoformat(),
    },
    "300750": {
        "code": "300750", "name": "宁德时代", "price": 182.50, "change": 1.86,
        "decision": "买入（信号一般）", "rule": "买入",
        "dif": 1.23, "dea": 0.85, "hist": 0.76,
        "bbU": 190, "bbM": 180, "bbL": 170, "bbPct": 0.63, "bbW": 11.0,
        "rsi": 55.3, "vr": 1.35, "tr": 2.14,
        "ma5": 180, "ma10": 178, "ma20": 180, "ma60": 175,
        "sups": [178, 175, 168], "ress": [190, 198, 210],
        "aiDecision": "买入", "aiConf": "中",
        "aiSum": "底部构筑完成，MACD金叉，量能温和放大，布林带开始上行扩口。整体技术形态偏多，建议分批介入。",
        "aiR": [
            "MACD零轴上方金叉，趋势转多",
            "量比1.35，稳健温和放量上涨",
            "布林中轨支撑有效，带口开始扩张"
        ],
        "aiE": 180, "aiS": 170, "aiT": 198,
        "macdD": [-1.2, -0.8, -0.3, 0.1, 0.4, 0.7, 0.9, 1.1, 1.2, 1.23],
        "deaD":  [-1.0, -0.9, -0.7, -0.5, -0.2, 0.1, 0.4, 0.65, 0.78, 0.85],
        "priceD": [172, 174, 176, 178, 179, 180, 181, 181.5, 182, 182.5],
        "lbls": ["5/20","5/21","5/22","5/23","5/24","5/27","5/28","5/29","5/30","5/31"],
        "updatedAt": datetime.datetime.now().isoformat(),
    },
}

# ══════════════════════════════════════════════════════════════
#  数据层
# ══════════════════════════════════════════════════════════════

# 内存存储，key = 股票代码
_stocks: dict = {}
_lock = threading.Lock()
_loading: set = set()   # 正在加载中的代码


def _init_stocks():
    """初始化：优先读 watchlist.json，否则读 .env，最后用 Demo 数据"""
    global _stocks
    if HAS_MODULES:
        try:
            cfg = Config()
            codes = _load_watchlist(cfg.stock_codes)
            _stocks = {code: {"code": code, "name": code, "loading": True} for code in codes}

            # 后台并发拉取，不阻塞 Flask 启动
            # 用 daemon 线程 + 错开启动时间，避免瞬间并发过高
            def _fetch_with_delay(idx: int, code: str):
                time.sleep(idx * 0.5)
                _fetch_stock(code)

            for i, code in enumerate(codes):
                t = threading.Thread(
                    target=_fetch_with_delay,
                    args=(i, code),
                    daemon=True,
                )
                t.start()
            return
        except Exception as e:
            print(f"[警告] 读取 Config 失败，使用 Demo 数据: {e}")
    _stocks = dict(DEMO_STOCKS)


def _pv(row, col, default=0.0):
    """安全取 DataFrame 行的数值"""
    v = row.get(col, default)
    try:
        fv = float(v)
        return fv if fv == fv else default  # NaN check
    except (TypeError, ValueError):
        return default


def _fetch_stock(code: str) -> dict:
    """拉取单只股票数据，写入 _stocks，返回数据 dict"""
    _loading.add(code)
    try:
        # 先拉股票基本信息（轻量，且失败时会返回 fallback {name: code}）
        # 拿到名字后立即更新 _stocks，让侧边栏先显示真实名字而不是裸代码
        info = fetch_stock_info(code)
        with _lock:
            existing = _stocks.get(code, {})
            existing.update({"code": code, "name": info.get("name", code), "loading": True})
            _stocks[code] = existing

        # 再拉 K 线（重型操作，可能失败）
        df = fetch_kline(code, days=60)
        if df is None or len(df) < 20:
            raise ValueError("数据不足")

        df = calc_indicators(df)
        sups, ress = find_support_resistance(df, keep=3)
        last = df.iloc[-1]
        prev = df.iloc[-2]

        price  = float(last["close"])
        change = float(last.get("change_amount", 0) or 0)

        # 支撑压力位关系
        sr_rel = describe_sr_relation(price, sups, ress)

        # 规则引擎
        rule_signal, rule_reasons = rule_engine(df, sups, ress, sr_rel)

        # AI（如已配置）
        ai_result = None
        try:
            cfg = Config()
            if cfg.enable_ai and cfg.ai_api_key:
                advisor = AIAdvisor(
                    api_key=cfg.ai_api_key,
                    base_url=cfg.ai_base_url,
                    model=cfg.ai_model,
                    temperature=cfg.ai_temperature,
                    max_tokens=cfg.ai_max_tokens,
                    timeout=cfg.ai_timeout,
                    enable_thinking=cfg.enable_thinking,
                    thinking_effort=cfg.thinking_effort,
                )
                prompt = build_prompt(
                    symbol=code,
                    stock_info=info,
                    df=df,
                    supports=sups,
                    resistances=ress,
                    sr_relation=sr_rel,
                )
                ai_result = advisor.query(prompt)
        except Exception as ae:
            print(f"[{code}] AI 分析跳过: {ae}")

        # 综合决策
        result = make_final_decision(
            symbol=code,
            stock_name=info.get("name", code),
            df=df,
            supports=sups,
            resistances=ress,
            sr_relation=sr_rel,
            ai_result=ai_result,
        )

        # 组装前端需要的数据格式
        stock_data = {
            "code": code,
            "name": info.get("name", code),
            "price": price,
            "change": change,
            "decision": result.final_decision,
            "rule": result.rule_signal,
            "ruleReasons": result.rule_reasons,

            # MACD
            "dif":  _pv(last, "dif"),
            "dea":  _pv(last, "dea"),
            "hist": _pv(last, "macd_hist"),

            # 布林带
            "bbU":  _pv(last, "bb_upper"),
            "bbM":  _pv(last, "bb_mid"),
            "bbL":  _pv(last, "bb_lower"),
            "bbPct": _pv(last, "bb_pct", 0.5),
            "bbW":  _pv(last, "bb_width"),

            # 其他指标
            "rsi": _pv(last, "rsi", 50),
            "vr":  _pv(last, "vol_ratio", 1),
            "tr":  _pv(last, "turnover_rate"),

            # 均线
            "ma5":  _pv(last, "ma5"),
            "ma10": _pv(last, "ma10"),
            "ma20": _pv(last, "ma20"),
            "ma60": _pv(last, "ma60"),

            # 支撑/压力
            "sups": sups or [round(price * 0.95, 2)],
            "ress": ress or [round(price * 1.05, 2)],

            # AI
            "aiDecision": result.ai_decision if result.ai_decision != "N/A" else rule_signal,
            "aiConf": result.ai_confidence if result.ai_confidence != "N/A" else "低",
            "aiSum":  result.ai_summary or "AI 未返回分析结果",
            "aiR":    result.ai_reasons or rule_reasons[1:4],
            "aiE":    result.ai_entry,
            "aiS":    result.ai_stop_loss,
            "aiTShort": result.ai_target_short,
            "aiTLong": result.ai_target_long,

            # 图表数据（最近10根，用于 MACD 图）
            "macdD":  df["dif"].tail(10).fillna(0).round(4).tolist(),
            "deaD":   df["dea"].tail(10).fillna(0).round(4).tolist(),
            "priceD": df["close"].tail(10).round(2).tolist(),
            "lbls":   df["date"].tail(10).dt.strftime("%m/%d").tolist(),

            # K 线蜡烛图数据（最近 60 根，ECharts candlestick 格式）
            # 每条：[日期, 开盘, 收盘, 最低, 最高, 成交量]
            "candles": [
                [
                    row["date"].strftime("%m/%d"),
                    round(float(row["open"]),   2),
                    round(float(row["close"]),  2),
                    round(float(row["low"]),    2),
                    round(float(row["high"]),   2),
                    int(row.get("volume", 0)),
                ]
                for _, row in df.tail(60).iterrows()
            ],

            # KDJ（最近 30 根）
            "kdjK": df["kdj_k"].tail(30).fillna(50).round(2).tolist() if "kdj_k" in df.columns else [],
            "kdjD": df["kdj_d"].tail(30).fillna(50).round(2).tolist() if "kdj_d" in df.columns else [],
            "kdjJ": df["kdj_j"].tail(30).fillna(50).round(2).tolist() if "kdj_j" in df.columns else [],
            "kdjLbls": df["date"].tail(30).dt.strftime("%m/%d").tolist(),

            # VWAP 最新值
            "vwap": _pv(last, "vwap"),

            # OBV 趋势（最近10根，判断量能）
            "obvD": df["obv"].tail(10).fillna(0).round(0).tolist() if "obv" in df.columns else [],

            "updatedAt": datetime.datetime.now().isoformat(),
            "loading": False,
            "error": None,
        }

        with _lock:
            _stocks[code] = stock_data

        # 检查价格预警
        _check_alerts(code, info.get("name", code), price)

        # WebSocket 实时推送：通知所有客户端这只股票已更新
        _ws_emit("stock_updated", stock_data)

        print(f"[{code}] ✓ 数据更新完成  价格={price}  决策={result.final_decision}")
        return stock_data

    except Exception as e:
        err = str(e)
        print(f"[{code}] ✗ 数据加载失败: {err}")
        with _lock:
            existing = _stocks.get(code, {})
            existing.update({"loading": False, "error": err})
            _stocks[code] = existing
        return _stocks.get(code, {})
    finally:
        _loading.discard(code)


# ══════════════════════════════════════════════════════════════
#  Flask App
# ══════════════════════════════════════════════════════════════

app = Flask(__name__, static_folder=".")

# ── WebSocket 实时推送（可选，未装 flask-socketio 时降级为纯 HTTP）────
# 推送事件：
#   stocks_snapshot       客户端连接时全量推
#   stock_updated         单只股票数据更新
#   commodities_snapshot  客户端连接时全量推 + 大宗商品刷新完成时
#   alert_triggered       价格预警命中时
try:
    from flask_socketio import SocketIO
    socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading",
                        logger=False, engineio_logger=False)
    HAS_WS = True
    print("[WS] flask-socketio 已加载")
except ImportError:
    socketio = None
    HAS_WS = False
    print("[WS] flask-socketio 未安装，将仅用 HTTP 模式")


def _ws_emit(event: str, payload):
    """安全推送：socketio 未启用时静默 no-op"""
    if not HAS_WS or socketio is None:
        return
    try:
        socketio.emit(event, payload)
    except Exception as e:
        # 推送失败不应影响主流程
        from loguru import logger
        logger.debug(f"[WS] emit {event} failed: {e}")


@app.after_request
def cors(r):
    r.headers["Access-Control-Allow-Origin"] = "*"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    r.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    return r


# ── 静态文件：dashboard ──────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(".", "dashboard.html")


# ── API: 全部股票列表（轻量，侧边栏用）────────────────────────
@app.route("/api/stocks")
def api_stocks():
    with _lock:
        codes = list(_stocks.keys())

    result = []
    for code in codes:
        with _lock:
            s = dict(_stocks.get(code, {}))
        # 只返回侧边栏需要的字段
        result.append({
            "code":     s.get("code", code),
            "name":     s.get("name", code),
            "price":    s.get("price", 0),
            "change":   s.get("change", 0),
            "decision": s.get("decision", "观望"),
            "rule":     s.get("rule", "观望"),
            "priceD":   s.get("priceD", []),
            "loading":  s.get("loading", False),
            "error":    s.get("error", None),
            "updatedAt": s.get("updatedAt", ""),
        })
    return jsonify({"ok": True, "stocks": result})


# ── API: 单只股票完整数据（主面板用）──────────────────────────
@app.route("/api/stock/<code>")
def api_stock(code):
    with _lock:
        s = dict(_stocks.get(code, {}))

    if not s:
        return jsonify({"ok": False, "error": f"股票 {code} 不存在"}), 404

    s.setdefault("loading", False)
    s.setdefault("error", None)
    return jsonify({"ok": True, "stock": s})


# ── API: 添加股票 ────────────────────────────────────────────
@app.route("/api/stock/add", methods=["POST", "OPTIONS"])
def api_add():
    if request.method == "OPTIONS":
        return "", 204
    data = request.get_json(force=True) or {}
    code = str(data.get("code", "")).strip()
    if not code:
        return jsonify({"ok": False, "error": "缺少 code 参数"}), 400

    with _lock:
        if code in _stocks:
            return jsonify({"ok": False, "error": f"{code} 已在监控列表中"}), 400

    # 先插入 loading 占位
    with _lock:
        _stocks[code] = {"code": code, "name": code, "loading": True, "error": None}

    if HAS_MODULES:
        t = threading.Thread(target=_fetch_stock, args=(code,), daemon=True)
        t.start()
    else:
        # Demo 模式
        with _lock:
            _stocks[code] = {
                **DEMO_STOCKS.get("000001", {}),
                "code": code, "name": code,
                "loading": False, "error": None,
                "updatedAt": datetime.datetime.now().isoformat(),
            }

    with _lock:
        _save_watchlist(list(_stocks.keys()))
    return jsonify({"ok": True, "code": code, "message": "已加入监控，数据加载中…"})


# ── API: 删除股票 ────────────────────────────────────────────
@app.route("/api/stock/remove", methods=["POST", "OPTIONS"])
def api_remove():
    if request.method == "OPTIONS":
        return "", 204
    data = request.get_json(force=True) or {}
    code = str(data.get("code", "")).strip()
    with _lock:
        if code not in _stocks:
            return jsonify({"ok": False, "error": "股票不存在"}), 404
        if len(_stocks) <= 1:
            return jsonify({"ok": False, "error": "至少保留一只股票"}), 400
        del _stocks[code]
    _save_watchlist([c for c in _stocks.keys()])
    return jsonify({"ok": True})


# ── API: 刷新单只数据 ────────────────────────────────────────
@app.route("/api/stock/refresh", methods=["POST", "OPTIONS"])
def api_refresh():
    if request.method == "OPTIONS":
        return "", 204
    data = request.get_json(force=True) or {}
    code = str(data.get("code", "")).strip()

    with _lock:
        if code not in _stocks:
            return jsonify({"ok": False, "error": "股票不存在"}), 404
        if code in _loading:
            return jsonify({"ok": False, "error": "正在加载中，请稍候"}), 429

    if HAS_MODULES:
        with _lock:
            _stocks[code]["loading"] = True
        t = threading.Thread(target=_fetch_stock, args=(code,), daemon=True)
        t.start()
        return jsonify({"ok": True, "message": "已开始刷新，约 5-10 秒后请求数据"})
    else:
        # Demo 模式：模拟小幅价格变动
        import random
        with _lock:
            s = _stocks.get(code, {})
            if s and "price" in s:
                s["price"] = round(s["price"] * (1 + random.uniform(-0.005, 0.005)), 2)
                s["updatedAt"] = datetime.datetime.now().isoformat()
                s["loading"] = False
        return jsonify({"ok": True, "message": "Demo 模式：价格已模拟刷新"})


# ── API: 服务状态 ────────────────────────────────────────────
@app.route("/api/status")
def api_status():
    return jsonify({
        "ok": True,
        "mode": "real" if HAS_MODULES else "demo",
        "commodity": HAS_COMMODITY,
        "commodity_ready": bool(_commodity_analysis),
        "commodity_loading": _analysis_loading.is_set(),
        "stocks": len(_stocks),
        "loading": list(_loading),
        "ws": HAS_WS,
        "time": datetime.datetime.now().isoformat(),
    })


# ══════════════════════════════════════════════════════════════
#  WebSocket 事件 handler
# ══════════════════════════════════════════════════════════════
if HAS_WS and socketio is not None:

    @socketio.on("connect")
    def _ws_on_connect():
        """客户端连接时，立即推送全量快照（股票 + 大宗商品）"""
        from flask import request
        sid = getattr(request, "sid", "?")
        print(f"[WS] 客户端连接 sid={sid}")
        try:
            with _lock:
                stocks_payload = list(_stocks.values())
            socketio.emit("stocks_snapshot", stocks_payload, to=sid)

            with _analysis_lock:
                comm_data = dict(_commodity_analysis)
            commodities = comm_data.pop("commodities", [])
            if commodities:
                socketio.emit("commodities_snapshot", {
                    "commodities": commodities,
                    "analysis":    comm_data,
                }, to=sid)
        except Exception as e:
            print(f"[WS] connect 推送失败: {e}")

    @socketio.on("disconnect")
    def _ws_on_disconnect():
        from flask import request
        sid = getattr(request, "sid", "?")
        print(f"[WS] 客户端断开 sid={sid}")


# ══════════════════════════════════════════════════════════════
#  大宗商品：后台抓取 + AI 分析
# ══════════════════════════════════════════════════════════════

def _get_commodity_advisor():
    """获取大宗商品 AI 分析器（无 key 时返回 None）"""
    if not HAS_COMMODITY:
        return None
    try:
        if not HAS_MODULES:
            return None
        cfg = Config()
        if not cfg.enable_ai or not cfg.ai_api_key:
            return None
        return CommodityAIAdvisor(
            api_key=cfg.ai_api_key,
            base_url=cfg.ai_base_url,
            model=cfg.ai_model,
            temperature=0.3,
            max_tokens=2000,
            timeout=90,
            enable_thinking=cfg.enable_thinking,
            thinking_effort=cfg.thinking_effort,
        )
    except Exception:
        return None


def _refresh_commodity_analysis():
    """后台线程：拉取大宗商品行情 + AI 分析"""
    if not HAS_COMMODITY:
        return
    if _analysis_loading.is_set():
        return
    _analysis_loading.set()
    try:
        commodities = commodity_fetcher.fetch_all()
        advisor = _get_commodity_advisor()
        if advisor:
            analysis = advisor.analyze(commodities)
        else:
            # 无 AI key 时使用规则引擎兜底
            analysis = CommodityAIAdvisor("", "", "").analyze(commodities)
        with _analysis_lock:
            _commodity_analysis.clear()
            _commodity_analysis.update(analysis)
            _commodity_analysis["commodities"] = commodities

        # 保存 AI 分析缓存
        _save_analysis_cache(dict(_commodity_analysis))

        # 更新价格历史（Sparkline 数据源）
        with _history_lock:
            for c in commodities:
                code = c["code"]
                if code not in _price_history:
                    _price_history[code] = deque(maxlen=24)
                _price_history[code].append(round(float(c["price"]), 4))

        # 检查大宗商品价格预警
        for c in commodities:
            _check_alerts(c["code"], c["name"], float(c["price"]))

        # WebSocket 实时推送：把整批大宗商品数据 + AI 分析结果一起推
        _ws_emit("commodities_snapshot", {
            "commodities": commodities,
            "analysis":    {k: v for k, v in _commodity_analysis.items() if k != "commodities"},
        })

        print(f"[大宗] ✓ {len(commodities)} 个品种，来源分布: "
              f"{set(c['source'] for c in commodities)}")
    except Exception as e:
        print(f"[大宗] ✗ {e}")
        import traceback; traceback.print_exc()
    finally:
        _analysis_loading.clear()


# ── API: 大宗商品实时行情 ────────────────────────────────────
@app.route("/api/commodities")
def api_commodities():
    if not HAS_COMMODITY:
        return jsonify({"ok": False, "error": "commodity_fetcher 模块未找到"}), 500
    try:
        data = commodity_fetcher.get_cached()
        if not data:
            threading.Thread(target=_refresh_commodity_analysis, daemon=True).start()
            return jsonify({"ok": True, "commodities": [], "loading": True})
        # 附加价格历史
        with _history_lock:
            for item in data:
                item["history"] = list(_price_history.get(item["code"], []))
        return jsonify({"ok": True, "commodities": data, "loading": False})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── API: AI 产业影响分析 ─────────────────────────────────────
@app.route("/api/commodity/analysis")
def api_commodity_analysis():
    if not HAS_COMMODITY:
        return jsonify({"ok": False, "error": "commodity_ai 模块未找到"}), 500
    with _analysis_lock:
        if not _commodity_analysis:
            threading.Thread(target=_refresh_commodity_analysis, daemon=True).start()
            return jsonify({
                "ok": True, "loading": True, "analysis": None,
                "message": "AI 分析生成中，约 15-30 秒后刷新…",
            })
        return jsonify({
            "ok": True,
            "loading": _analysis_loading.is_set(),
            "analysis": dict(_commodity_analysis),
        })


# ── API: 强制刷新大宗商品 ────────────────────────────────────
@app.route("/api/commodity/refresh", methods=["POST", "OPTIONS"])
def api_commodity_refresh():
    if request.method == "OPTIONS":
        return "", 204
    if not HAS_COMMODITY:
        return jsonify({"ok": False, "error": "commodity_fetcher 模块未找到"}), 500
    if _analysis_loading.is_set():
        return jsonify({"ok": False, "message": "正在刷新中，请稍候"}), 429
    commodity_fetcher._last_fetch = 0   # 清除缓存，强制重新拉取
    threading.Thread(target=_refresh_commodity_analysis, daemon=True).start()
    return jsonify({"ok": True, "message": "已触发刷新"})


# ── API: 板块联动分析 ────────────────────────────────────────
@app.route("/api/stock/<code>/sector")
def api_sector(code):
    if not HAS_EXTRA:
        return jsonify({"ok": False, "error": "sector_fetcher 模块未找到"}), 500
    try:
        data = get_sector_info(code)
        return jsonify({"ok": True, "sector": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── API: 北向资金 ────────────────────────────────────────────
@app.route("/api/market/northbound")
def api_northbound():
    if not HAS_EXTRA:
        return jsonify({"ok": False, "error": "capital_fetcher 模块未找到"}), 500
    try:
        data = get_northbound()
        return jsonify({"ok": True, "northbound": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── API: 个股融资融券 ─────────────────────────────────────────
@app.route("/api/stock/<code>/margin")
def api_margin(code):
    if not HAS_EXTRA:
        return jsonify({"ok": False, "error": "capital_fetcher 模块未找到"}), 500
    try:
        data = get_margin(code)
        return jsonify({"ok": True, "margin": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── API: 财报日历 ────────────────────────────────────────────
@app.route("/api/calendar")
def api_calendar():
    if not HAS_EXTRA:
        return jsonify({"ok": False, "error": "calendar_fetcher 模块未找到"}), 500
    try:
        with _lock:
            codes = list(_stocks.keys())
        events = get_upcoming_events(codes)
        return jsonify({"ok": True, "events": events})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  新功能 API
# ══════════════════════════════════════════════════════════════

# ── 价格预警 ─────────────────────────────────────────────────
@app.route("/api/alerts", methods=["GET"])
def api_alerts_get():
    with _alerts_lock:
        return jsonify({"ok": True, "alerts": list(_alerts)})

@app.route("/api/alerts", methods=["POST", "OPTIONS"])
def api_alerts_add():
    if request.method == "OPTIONS":
        return "", 204
    data = request.get_json(force=True) or {}
    code      = str(data.get("code", "")).strip()
    name      = str(data.get("name", code))
    direction = str(data.get("direction", "above"))   # above / below
    threshold = float(data.get("threshold", 0))
    price_type = str(data.get("type", "stock"))       # stock / commodity
    if not code or not threshold:
        return jsonify({"ok": False, "error": "缺少 code 或 threshold"}), 400
    alert = {
        "id":        f"{code}_{direction}_{threshold}_{int(time.time())}",
        "type":      price_type,
        "code":      code,
        "name":      name,
        "direction": direction,
        "threshold": threshold,
        "triggered": False,
        "triggered_at": None,
        "created_at": datetime.datetime.now().isoformat(),
    }
    with _alerts_lock:
        _alerts.append(alert)
    _save_alerts()
    return jsonify({"ok": True, "alert": alert})

@app.route("/api/alerts/<alert_id>", methods=["DELETE", "OPTIONS"])
def api_alerts_delete(alert_id):
    if request.method == "OPTIONS":
        return "", 204
    with _alerts_lock:
        before = len(_alerts)
        _alerts[:] = [a for a in _alerts if a.get("id") != alert_id]
        deleted = len(_alerts) < before
    _save_alerts()
    return jsonify({"ok": deleted})

@app.route("/api/alerts/triggered")
def api_alerts_triggered():
    """返回已触发的预警并清空队列"""
    items = list(_triggered_queue)
    _triggered_queue.clear()
    return jsonify({"ok": True, "triggered": items})

@app.route("/api/alerts/reset/<alert_id>", methods=["POST", "OPTIONS"])
def api_alerts_reset(alert_id):
    """重置预警（让它可以再次触发）"""
    if request.method == "OPTIONS":
        return "", 204
    with _alerts_lock:
        for a in _alerts:
            if a.get("id") == alert_id:
                a["triggered"] = False
                a["triggered_at"] = None
    _save_alerts()
    return jsonify({"ok": True})


# ── 每日市场总结（按钮触发）────────────────────────────────
@app.route("/api/market/summary", methods=["POST", "OPTIONS"])
def api_market_summary():
    if request.method == "OPTIONS":
        return "", 204
    if not HAS_MODULES:
        return jsonify({"ok": False, "error": "Demo 模式不支持 AI 总结"}), 400
    try:
        cfg = Config()
        if not cfg.enable_ai or not cfg.ai_api_key:
            return jsonify({"ok": False, "error": "未配置 AI Key，请在 .env 中设置 DEEPSEEK_API_KEY"}), 400

        from ai_advisor import AIAdvisor
        from openai import OpenAI

        # 收集股票数据
        with _lock:
            stocks_snap = list(_stocks.values())

        # 收集大宗商品数据
        with _analysis_lock:
            comm_snap = _commodity_analysis.get("commodities", [])

        # 构建综合 Prompt
        today = datetime.datetime.now().strftime("%Y年%m月%d日")
        stock_lines = []
        for s in stocks_snap:
            if s.get("loading") or not s.get("price"):
                continue
            stock_lines.append(
                f"  {s.get('name',s['code'])}({s['code']}): "
                f"价格{s.get('price',0)} 涨跌{s.get('change',0):+.2f} "
                f"决策:{s.get('decision','观望')}"
            )
        comm_lines = []
        for c in comm_snap:
            pct = c.get("change_pct", 0)
            comm_lines.append(
                f"  {c['name']}({c['code']}): {c['price']} {c.get('unit','')}  {pct:+.2f}%"
            )

        prompt = f"""请为以下{today}的市场数据生成一份简洁的每日市场总结报告。

【监控股票】
{chr(10).join(stock_lines) if stock_lines else "  暂无数据"}

【全球大宗商品 & 外汇】
{chr(10).join(comm_lines) if comm_lines else "  暂无数据"}

要求：
1. 整体市场情绪判断（乐观/中性/谨慎/悲观）
2. 今日最值得关注的 2-3 个信号（股票或大宗商品）
3. 大宗商品价格变动对A股的潜在影响（1-2句）
4. 明日操作建议（1-2句）
5. 字数控制在 200 字以内，语言简洁专业

直接输出中文正文，不需要 JSON，不需要标题。"""

        client = OpenAI(api_key=cfg.ai_api_key, base_url=cfg.ai_base_url, timeout=60)
        kwargs = dict(
            model=cfg.ai_model,
            messages=[
                {"role": "system", "content": "你是资深A股策略分析师，擅长简洁、准确地总结市场每日动态。"},
                {"role": "user", "content": prompt},
            ],
            max_tokens=600,
        )
        if cfg.enable_thinking:
            kwargs["reasoning_effort"] = cfg.thinking_effort
            kwargs["extra_body"] = {"thinking": {"type": "enabled"}}
        else:
            kwargs["temperature"] = cfg.ai_temperature

        resp = client.chat.completions.create(**kwargs)
        summary = resp.choices[0].message.content.strip()

        return jsonify({
            "ok": True,
            "summary": summary,
            "generated_at": datetime.datetime.now().isoformat(),
            "date": today,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


# ── 数据导出 Excel（按钮触发）────────────────────────────────
@app.route("/api/export")
def api_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from io import BytesIO
        from flask import send_file

        wb = openpyxl.Workbook()

        # ── Sheet 1: 股票 ────────────────────────────────────
        ws1 = wb.active
        ws1.title = "股票监控"
        headers1 = ["代码","名称","价格","涨跌额","决策","AI建议",
                    "RSI","量比","MA5","MA20","DIF","DEA","更新时间"]
        ws1.append(headers1)
        with _lock:
            stocks_data = list(_stocks.values())
        for s in stocks_data:
            if s.get("loading"):
                continue
            ws1.append([
                s.get("code",""), s.get("name",""),
                s.get("price",0), s.get("change",0),
                s.get("decision",""), s.get("aiDecision",""),
                s.get("rsi",0), s.get("vr",0),
                s.get("ma5",0), s.get("ma20",0),
                s.get("dif",0), s.get("dea",0),
                (s.get("updatedAt","") or "")[:19],
            ])

        # ── Sheet 2: 大宗商品 & 外汇 ──────────────────────
        ws2 = wb.create_sheet("大宗商品&外汇")
        headers2 = ["代码","名称","分类","价格","单位","涨跌额","涨跌幅%","数据来源","更新时间"]
        ws2.append(headers2)
        with _analysis_lock:
            comm_data = _commodity_analysis.get("commodities", [])
        for c in comm_data:
            ws2.append([
                c.get("code",""), c.get("name",""), c.get("category",""),
                c.get("price",0), c.get("unit",""),
                c.get("change",0), c.get("change_pct",0),
                c.get("source",""), (c.get("updated_at","") or "")[:19],
            ])

        # ── 样式：加粗表头 ────────────────────────────────
        for ws in [ws1, ws2]:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="1E2228")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 14

        # ── 输出为字节流 ──────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"StockMonitor_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True,
                         download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl 未安装，请运行 pip install openpyxl"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  自动定时刷新（仅真实模式）
# ══════════════════════════════════════════════════════════════

def _auto_refresh_loop():
    if not HAS_MODULES:
        return
    try:
        cfg = Config()
        stock_interval = cfg.interval_minutes * 60
    except Exception:
        stock_interval = 3600

    commodity_interval = 300   # 大宗商品 5 分钟刷新一次
    stock_ts = commodity_ts = time.time()

    while True:
        time.sleep(60)
        now = time.time()

        # 大宗商品定时刷新
        if HAS_COMMODITY and (now - commodity_ts) >= commodity_interval:
            commodity_ts = now
            threading.Thread(target=_refresh_commodity_analysis, daemon=True).start()

        # 股票定时刷新
        if now - stock_ts >= stock_interval:
            stock_ts = now
            print(f"[定时] 开始自动刷新所有股票 ({len(_stocks)} 只)...")
            with _lock:
                codes = list(_stocks.keys())
            for code in codes:
                t = threading.Thread(target=_fetch_stock, args=(code,), daemon=True)
                t.start()
                time.sleep(2)


# ══════════════════════════════════════════════════════════════
#  主入口
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 55)
    print("  StockMonitor · Web 仪表盘后端")
    print("=" * 55)
    print(f"  模式：{'真实数据 (akshare + AI)' if HAS_MODULES else 'Demo 演示数据'}")
    print("  地址：http://localhost:5000")
    print("  按 Ctrl+C 停止")
    print("=" * 55)

    # 启动前先探测数据源健康状况（约 10 秒）。这样后续每只股票直接走
    # 可用源，不必再"先试东财再 fallback"，省时间也减少骚扰对端
    if HAS_MODULES:
        try:
            from fetcher import probe_data_sources
            probe_data_sources()
        except Exception as e:
            print(f"[源探测] 跳过（{e}）")

    _init_stocks()

    # 启动时加载缓存和预警
    _load_alerts()
    cached = _load_analysis_cache()
    if cached and HAS_COMMODITY:
        with _analysis_lock:
            _commodity_analysis.update(cached)
        print("[缓存] 已从缓存恢复上次 AI 分析结果")

    # 启动时异步加载大宗商品（即使有缓存也后台刷新）
    if HAS_COMMODITY:
        threading.Thread(target=_refresh_commodity_analysis, daemon=True).start()

    # 后台自动刷新线程
    t = threading.Thread(target=_auto_refresh_loop, daemon=True)
    t.start()

    # 启动 web 服务：有 ws 时用 socketio.run，否则用普通 app.run
    if HAS_WS and socketio is not None:
        print("[启动] 使用 SocketIO 模式（HTTP + WebSocket）")
        socketio.run(app, host="0.0.0.0", port=5000, debug=False,
                     allow_unsafe_werkzeug=True)
    else:
        print("[启动] 使用纯 HTTP 模式")
        app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
